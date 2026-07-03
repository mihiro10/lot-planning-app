"""
Parse a planning Excel file and return structured product + daily-value data.

Column positions are NOT hardcoded — different sheets (e.g. "新計画6月その2.xlsx"
vs "昼計画6月") have reordered columns (分類 sits at column D in one, column J in
the other). Instead, row 2's header text is scanned to locate each field by
label, so the parser adapts to whichever layout a given sheet actually uses
instead of silently reading the wrong column into the wrong field.
"""
from datetime import datetime, date, timedelta
from io import BytesIO
import openpyxl

# field -> header substrings to match, in priority order (first match wins).
# Longer/more specific substrings should come before shorter ones that could
# also match a different field's header.
HEADER_PATTERNS = {
    "row_type":            ["行種別"],
    "name":                ["品目名"],
    "code":                ["コード"],
    "planner":             ["計画担当"],
    "mfg_location":        ["製造場所"],
    "storage":             ["最終入庫場所", "入庫場所"],
    "category":            ["分類"],
    "mfg_name":             ["製造名"],
    "lot_unit":            ["ロット単位"],
    "lot_size":            ["ロット"],
    "min_stock":           ["最低在庫"],
    "max_stock":           ["最高在庫"],
    "lead_time":           ["入庫リードタイム", "リードタイム"],
    "starting_inventory":  ["在庫数"],
    "pre_inventory":       ["棚卸し前"],
    "notes":               ["備考"],
}

MANUAL_ROLES = {
    "計画（倍）":  "plan",
    "入庫":        "inbound_actual",
    "使用予測":    "demand_forecast",
    "店":          "demand_actual",
    "通販":        "demand_actual",
    "米飯課":      "demand_actual",
    "調整":        "adjustment",
}


def _parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _detect_columns(ws, header_row: int = 2) -> dict:
    """Scan the header row and return {field_name: column_index} by matching
    header text. Raises if row_type or name (the two anchors everything else
    depends on) aren't found."""
    header_cells = [(c, ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    header_cells = [(c, str(v)) for c, v in header_cells if v is not None]

    cols = {}
    used = set()
    for field, patterns in HEADER_PATTERNS.items():
        for pattern in patterns:
            match = next((c for c, text in header_cells if pattern in text and c not in used), None)
            if match:
                cols[field] = match
                used.add(match)
                break

    if "row_type" not in cols:
        raise ValueError("「行種別」列が見つかりません。列レイアウトを確認してください。")
    if "name" not in cols:
        raise ValueError("「品目名」列が見つかりません。列レイアウトを確認してください。")

    cols["day1"] = cols["row_type"] + 1
    return cols


def parse_excel(file_bytes: bytes) -> dict:
    """
    Returns:
      {
        "start_date": "2026-06-01",
        "end_date":   "2026-06-30",
        "products": [
          {
            "code": ..., "name": ..., "planner": ..., ...,
            "starting_inventory": ..., "pre_inventory": ...,
            "daily_values": {"計画（倍）": {"2026-06-01": 1, ...}, ...},
            "daily_notes":  {"2026-06-01": "追加", ...},
          }
        ]
      }
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    col = _detect_columns(ws)
    col_rtype = col["row_type"]
    col_day1 = col["day1"]

    # Only the day1 header cell is a real date value; subsequent cells are
    # formulas (=T2+1 etc.) whose cached values may be absent — derive end
    # from max_column instead.
    start_date = _parse_date(ws.cell(2, col_day1).value)
    if not start_date:
        raise ValueError("日付ヘッダーが見つかりません。列レイアウトを確認してください。")

    num_day_cols = ws.max_column - col_day1
    end_date = start_date + timedelta(days=num_day_cols)

    dates = []
    d = start_date
    while d <= end_date:
        dates.append(d)
        d += timedelta(days=1)

    def _v(row, field):
        c = col.get(field)
        return ws.cell(row, c).value if c else None

    def _float(row, field, default=0.0):
        try:    return float(_v(row, field))
        except: return default

    def _int(row, field, default=0):
        try:    return int(_v(row, field))
        except: return default

    products = []
    current = None

    for row in range(3, ws.max_row + 1):
        rtype = ws.cell(row, col_rtype).value
        if not rtype:
            continue

        if rtype == "備考":
            name = _v(row, "name")
            if not name:
                current = None
                continue

            # Skip template rows: name is literally the template placeholder, or
            # lot_size holds leftover header text ("ロット内数") instead of a number.
            # NOTE: a real product's code cell can legitimately be blank or contain
            # placeholder text like "code" — that alone must NOT skip the row, or
            # genuine products (e.g. きゅうりの醤油漬) get silently dropped.
            raw_code = _v(row, "code")
            lot_size_raw = _v(row, "lot_size")
            if "テンプレート" in str(name) or (lot_size_raw is not None and not isinstance(lot_size_raw, (int, float))):
                current = None
                continue

            current = {
                "code":                str(raw_code) if raw_code else None,
                "name":                name,
                "planner":             _v(row, "planner"),
                "mfg_location":        _v(row, "mfg_location"),
                "category":            _v(row, "category"),
                "storage":             _v(row, "storage"),
                "mfg_name":            _v(row, "mfg_name"),
                "lot_size":            _float(row, "lot_size", 1.0),
                "lot_unit":            str(_v(row, "lot_unit")) if _v(row, "lot_unit") else "p",
                "min_stock":           _float(row, "min_stock"),
                "max_stock":           _float(row, "max_stock"),
                "lead_time":           _int(row, "lead_time"),
                "notes":               _v(row, "notes"),
                "starting_inventory":  _float(row, "starting_inventory"),
                "pre_inventory":       _float(row, "pre_inventory"),
                "daily_values":        {},
                "daily_notes":         {},
            }
            # The 備考 row's own day-columns carry free-text annotations per date
            # (e.g. "追加", "変更70ｐ", "キャンセル") — these are what floor workers
            # actually read, distinct from the header 備考 note field above.
            for i, d in enumerate(dates):
                text = ws.cell(row, col_day1 + i).value
                if text not in (None, ""):
                    current["daily_notes"][d.isoformat()] = str(text)
            products.append(current)
            continue

        if current is None:
            continue

        if rtype in MANUAL_ROLES:
            day_vals = {}
            for i, d in enumerate(dates):
                val = ws.cell(row, col_day1 + i).value
                if val is not None:
                    try:
                        day_vals[d.isoformat()] = float(val)
                    except (TypeError, ValueError):
                        pass
            if day_vals:
                # Multiple rows may share same rtype name (e.g. two demand_actual rows)
                current["daily_values"][rtype] = day_vals

    return {
        "start_date": start_date.isoformat(),
        "end_date":   end_date.isoformat(),
        "products":   products,
    }
